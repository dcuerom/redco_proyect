"""Carga y deriva todas las cifras de `pilar_a` que alimentan la presentación.

Fuente única: `pilar_a/deliverables/REDCO_Rentabilidad_Ventas_FlujoCaja_historico.xlsx`
(el libro de trazabilidad) más el comparativo de dotación 2025-2026.

Convenciones fijadas tras verificación (ver plan):
  * VENTA = 'Emitido USD' del ledger CicloEdP, atribuido por el campo `Anio`
    (año de proyecto), NO por `Anio_Emision`. Solo `Anio` reproduce los totales
    de la hoja 02_Ventas (2023=7,071,137 · 2024=8,418,845 · 2025=9,476,037).
  * Cliente = `Cliente_Cons` (nombre consolidado; normaliza ElBrocal→El Brocal, etc.).
  * Perímetro de rentabilidad = 4 países (Chile+Perú+Brasil+USA), igual que rev1 s.4,
    excluyendo 'Otros'/BVI/Rusia. Se declara en el pie de cada slide.
"""

import collections
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
XL_HIST = ROOT / "pilar_a/deliverables/REDCO_Rentabilidad_Ventas_FlujoCaja_historico.xlsx"
XL_DOT = ROOT / "pilar_a/data/Archivos 2026/Dotación_Comparativo_Estandarizado_2025-2026.xlsx"

PAISES_CORE = ["Chile", "Perú", "Brasil", "USA"]


def _sheet_rows(wb, name, header_row):
    ws = wb[name]
    hdr = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    idx = {h: i for i, h in enumerate(hdr) if h}
    data = [r for r in ws.iter_rows(min_row=header_row + 1, values_only=True)
            if any(c is not None for c in r)]
    return idx, data


def _eff_n(values):
    """Número efectivo de participantes = 1/HHI."""
    tot = sum(v for v in values if v and v > 0)
    if not tot:
        return 0.0
    return 1.0 / sum((v / tot) ** 2 for v in values if v and v > 0)


def cargar():
    wb = openpyxl.load_workbook(XL_HIST, read_only=True, data_only=True)
    d = {}

    # ---- ledger EdP -------------------------------------------------------
    idx, led = _sheet_rows(wb, "90_Origen_Ledger_EdP", 2)
    iA, iP, iE = idx["Anio"], idx["Pais"], idx["Emitido_USD"]
    iC, iD = idx["Cliente_Cons"], idx["Dias_Emision_Ingreso"]

    venta_pais = collections.defaultdict(float)
    venta_cli = collections.defaultdict(float)
    ciclo = collections.defaultdict(list)
    for r in led:
        anio, emit = r[iA], (r[iE] or 0)
        if anio is not None and emit > 0:
            venta_pais[(int(anio), r[iP])] += emit
            venta_cli[(int(anio), r[iC] or r[iP])] += emit
        dias = r[iD]
        if isinstance(dias, (int, float)) and dias >= 0:
            ciclo[r[iP]].append(dias)

    anios = [2023, 2024, 2025, 2026]
    d["venta_pais"] = {
        p: [round(venta_pais.get((a, p), 0.0)) for a in anios]
        for p in ["Chile", "Perú", "Brasil", "USA", "Rusia", "BVI", "Canadá"]
    }
    d["venta_total"] = [round(sum(v for (y, _), v in venta_pais.items() if y == a)) for a in anios]
    d["anios_venta"] = anios
    d["paises_efectivos"] = {
        a: round(_eff_n([v for (y, _), v in venta_pais.items() if y == a]), 2) for a in anios
    }

    d["top_clientes"] = {}
    for a in (2023, 2024, 2025):
        items = sorted(((c, v) for (y, c), v in venta_cli.items() if y == a), key=lambda x: -x[1])
        tot = sum(v for _, v in items)
        top4 = [(c, round(v), round(v / tot * 100, 1)) for c, v in items[:4]]
        resto = sum(v for _, v in items[4:])
        d["top_clientes"][a] = {
            "total": round(tot),
            "n_clientes": len(items),
            "clientes_efectivos": round(_eff_n([v for _, v in items]), 1),
            "top4": top4,
            "resto": (f"Otros ({len(items) - 4})", round(resto), round(resto / tot * 100, 1)),
            "top1_pct": round(top4[0][2]),
            "top3_pct": round(sum(t[2] for t in top4[:3])),
        }

    d["ciclo_pais"] = {p: round(sum(v) / len(v), 1) for p, v in ciclo.items() if v}
    todos = [v for L in ciclo.values() for v in L]
    d["ciclo_grupo"] = round(sum(todos) / len(todos), 1)
    d["rotacion_grupo"] = round(365 / d["ciclo_grupo"], 1)
    d["ciclo_n"] = {p: len(v) for p, v in ciclo.items()}

    # ---- rentabilidad 2024 por país --------------------------------------
    idx, rows24 = _sheet_rows(wb, "90_Origen_Margen2024_Pais", 2)
    rent = {}
    for r in rows24:
        pais = r[idx["Pais"]]
        if pais:
            rent[pais] = {"ing_2024": r[idx["Ingreso_USD"]], "gasto_2024": r[idx["Gasto_USD"]]}

    # ---- rentabilidad 2025 por país --------------------------------------
    idx, rows25 = _sheet_rows(wb, "90_Origen_Rentab_2025", 2)
    acc = collections.defaultdict(lambda: collections.defaultdict(float))
    for r in rows25:
        pais, concepto, val = r[idx["Pais"]], r[idx["Concepto"]], r[idx["Valor_USD"]]
        if pais and concepto:
            acc[pais][concepto] += (val or 0)
    for pais, c in acc.items():
        rent.setdefault(pais, {})
        rent[pais]["ing_2025"] = c.get("INGRESOS OP", 0.0)
        rent[pais]["gasto_2025"] = c.get("COSTOS OP", 0.0)

    for pais, v in rent.items():
        for a in ("2024", "2025"):
            ing, gasto = v.get(f"ing_{a}"), v.get(f"gasto_{a}")
            if ing:
                v[f"benef_{a}"] = ing - gasto
                v[f"margen_{a}"] = (ing - gasto) / ing
    d["rentabilidad"] = rent

    # consolidado de 4 países (perímetro rev1 s.4)
    d["consolidado4"] = {}
    for a in ("2024", "2025"):
        ing = sum(rent[p].get(f"ing_{a}", 0) for p in PAISES_CORE)
        gasto = sum(rent[p].get(f"gasto_{a}", 0) for p in PAISES_CORE)
        d["consolidado4"][a] = {"ingreso": ing, "gasto": gasto, "margen_usd": ing - gasto,
                                "margen_pct": (ing - gasto) / ing}
    # consolidado total (incluye Otros/Rusia) — para las metas
    d["consolidado_total"] = {
        "2024": {"ingreso": 8274848.2, "benef": 2307005.1, "margen_pct": 2307005.1 / 8274848.2},
        "2025": {"ingreso": 9150616.3, "benef": 1431778.8, "margen_pct": 1431778.8 / 9150616.3},
    }

    # ---- rentabilidad por nivel de estudio (hoja 06) -----------------------
    ws = wb["06_Rentab_Nivel_Modulos"]
    niveles = []
    for r in ws.iter_rows(min_row=1, values_only=True):
        if r[0] and isinstance(r[2], (int, float)) and isinstance(r[5], (int, float)):
            if str(r[0]).startswith("TOTAL"):
                d["nivel_total"] = {"n": r[1], "ingreso": r[2], "costo": r[3],
                                    "rent": r[4], "margen": r[5]}
            else:
                niveles.append({"nivel": r[0], "n": r[1], "ingreso": r[2],
                                "costo": r[3], "rent": r[4], "margen": r[5]})
    d["niveles"] = sorted(niveles, key=lambda x: -x["margen"])

    # ---- KPI de propuestas (conversión y ticket) --------------------------
    idx, kpi = _sheet_rows(wb, "90_Origen_KPI_Propuestas", 2)
    agg = collections.defaultdict(lambda: collections.defaultdict(float))
    for r in kpi:
        a = r[idx["Anio"]]
        if a is None:
            continue
        for k in ("Emitidas_num", "Emitidas_USD", "Adjudicadas_num", "Adjudicadas_USD"):
            agg[int(a)][k] += (r[idx[k]] or 0)
    d["propuestas"] = {}
    for a, v in sorted(agg.items()):
        if not v["Emitidas_USD"]:
            continue
        d["propuestas"][a] = {
            "emit_n": int(v["Emitidas_num"]), "emit_usd": v["Emitidas_USD"],
            "adj_n": int(v["Adjudicadas_num"]), "adj_usd": v["Adjudicadas_USD"],
            "conv_valor": v["Adjudicadas_USD"] / v["Emitidas_USD"],
            "conv_num": v["Adjudicadas_num"] / v["Emitidas_num"] if v["Emitidas_num"] else 0,
            "ticket_emit": v["Emitidas_USD"] / v["Emitidas_num"] if v["Emitidas_num"] else 0,
            "ticket_adj": v["Adjudicadas_USD"] / v["Adjudicadas_num"] if v["Adjudicadas_num"] else 0,
        }

    # ---- fase FEL: Core vs Adyacencia (hoja 01 bloque C) ------------------
    ws = wb["01_Rentabilidad_2024_2025"]
    fel = {}
    for r in ws.iter_rows(min_row=1, values_only=True):
        if r[0] and str(r[0]).startswith("Subtotal") and isinstance(r[4], (int, float)):
            fel[str(r[0]).replace("Subtotal ", "").strip()] = r[4]
        if r[0] and str(r[0]).startswith("TOTAL Emitido") and isinstance(r[4], (int, float)):
            fel["TOTAL"] = r[4]
    d["fel"] = fel
    wb.close()

    # ---- dotación 2025 vs 2026 -------------------------------------------
    wbd = openpyxl.load_workbook(XL_DOT, read_only=True, data_only=True)
    ws = wbd["Comparativo por País"]
    filas = [r for r in ws.iter_rows(min_row=6, values_only=True) if r[0]]
    cats, dot = [], {}
    for r in filas:
        if str(r[0]).startswith("Total"):
            dot["TOTAL"] = {"2025": r[11], "2026": r[12]}
        else:
            cats.append(r[0])
            dot[r[0]] = {"2025": r[11], "2026": r[12]}
    d["dotacion_cat"] = dot
    d["dotacion_cats"] = cats
    d["dotacion_pais"] = {}
    for i, p in enumerate(["Brasil", "Chile", "Perú", "Rusia", "USA"]):
        tot_row = [r for r in filas if str(r[0]).startswith("Total")][0]
        d["dotacion_pais"][p] = {"2025": tot_row[1 + i * 2], "2026": tot_row[2 + i * 2]}
    wbd.close()

    # productividad: ingreso operacional / dotación
    d["ingreso_x_persona"] = {
        2025: d["consolidado_total"]["2025"]["ingreso"] / dot["TOTAL"]["2025"],
        2026: d["consolidado_total"]["2025"]["ingreso"] / dot["TOTAL"]["2026"],
    }
    return d


if __name__ == "__main__":
    import json
    datos = cargar()
    print(json.dumps(datos, indent=2, ensure_ascii=False, default=float))
